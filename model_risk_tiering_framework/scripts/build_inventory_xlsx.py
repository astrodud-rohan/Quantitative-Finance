"""
Builds the Model_Inventory.xlsx deliverable with TWO sheets:
    - Model Inventory - 18-model inventory with live formulas for scoring and tiering
    - Tiering Summary - summary of model counts by tier and risk type
"""
import sys
sys.path.append("src")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from build_inventory import build_inventory_df

ARIAL = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=ARIAL, size=16, bold=True, color="1F3864")
SUBTITLE_FONT = Font(name=ARIAL, size=10, italic=True, color="888888")
BODY_FONT = Font(name=ARIAL, size=10)
INPUT_FONT = Font(name=ARIAL, size=10, color="0000FF")
FORMULA_FONT = Font(name=ARIAL, size=10, color="000000")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TIER1_FILL = PatternFill("solid", fgColor="FCE4E4")
TIER2_FILL = PatternFill("solid", fgColor="FFF3CD")
TIER3_FILL = PatternFill("solid", fgColor="E2EFDA")

def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def build_workbook():
    df = build_inventory_df()
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Inventory"
 
    # ---- Title block ----
    ws["A1"] = "Model Risk Inventory and Tiering"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "As of: FY2026 Annual Review  |  Owner: Model Risk Management"
    ws["A2"].font = SUBTITLE_FONT
 
    # ---- Weights reference block (top-right, inputs in blue) ----
    weight_start_col = 18  # column R, safely past the 15-column data table (ends at col O)
    ws.cell(row=1, column=weight_start_col, value="Tiering Weights (editable inputs)").font = Font(name=ARIAL, bold=True, size=10)
    weight_labels = [("Materiality", 0.40), ("Complexity", 0.25), ("Usage/Reliance", 0.25), ("Data Quality Risk", 0.10)]
    for i, (label, val) in enumerate(weight_labels):
        r = 2 + i
        ws.cell(row=r, column=weight_start_col, value=label).font = BODY_FONT
        cell = ws.cell(row=r, column=weight_start_col + 1, value=val)
        cell.font = INPUT_FONT
        cell.number_format = "0%"
    weight_cells = {
        "materiality": f"${get_column_letter(weight_start_col+1)}$2",
        "complexity": f"${get_column_letter(weight_start_col+1)}$3",
        "usage_reliance": f"${get_column_letter(weight_start_col+1)}$4",
        "data_quality_risk": f"${get_column_letter(weight_start_col+1)}$5",
    }
 
    ws.cell(row=7, column=weight_start_col, value="Tier Thresholds (editable inputs)").font = Font(name=ARIAL, bold=True, size=10)
    threshold_labels = [("Tier 1 (High) >=", 3.60), ("Tier 2 (Medium) >=", 2.40)]
    for i, (label, val) in enumerate(threshold_labels):
        r = 8 + i
        ws.cell(row=r, column=weight_start_col, value=label).font = BODY_FONT
        cell = ws.cell(row=r, column=weight_start_col + 1, value=val)
        cell.font = INPUT_FONT
    tier1_threshold_cell = f"${get_column_letter(weight_start_col+1)}$8"
    tier2_threshold_cell = f"${get_column_letter(weight_start_col+1)}$9"
 
    # ---- Table header ----
    header_row = 4
    headers = ["Model ID", "Model Name", "Model Type", "Owner Function", "Business Line",
               "Materiality\n(1-5)", "Complexity\n(1-5)", "Usage/\nReliance (1-5)",
               "Data Quality\nRisk (1-5)", "Weighted\nScore", "Tier",
               "Validation\nFrequency", "Monitoring\nFrequency", "Sign-off\nRequired", "Scoring Rationale"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(headers))
 
    # ---- Data rows with live formulas ----
    from tiering_engine import VALIDATION_REQUIREMENTS
 
    for i, row_data in df.iterrows():
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=row_data["model_id"]).font = BODY_FONT
        ws.cell(row=r, column=2, value=row_data["model_name"]).font = BODY_FONT
        ws.cell(row=r, column=3, value=row_data["model_type"]).font = BODY_FONT
        ws.cell(row=r, column=4, value=row_data["owner_function"]).font = BODY_FONT
        ws.cell(row=r, column=5, value=row_data["business_line"]).font = BODY_FONT
        for c in range(1, 6):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center", wrap_text=True)
 
        for j, dim in enumerate(["materiality", "complexity", "usage_reliance", "data_quality_risk"]):
            cell = ws.cell(row=r, column=6 + j, value=int(row_data[dim]))
            cell.font = INPUT_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
 
        # Weighted score formula referencing the weight cells, not hardcoded weights
        score_formula = (
            f"=F{r}*{weight_cells['materiality']}+G{r}*{weight_cells['complexity']}"
            f"+H{r}*{weight_cells['usage_reliance']}+I{r}*{weight_cells['data_quality_risk']}"
        )
        score_cell = ws.cell(row=r, column=10, value=score_formula)
        score_cell.font = FORMULA_FONT
        score_cell.number_format = "0.00"
        score_cell.alignment = Alignment(horizontal="center", vertical="center")
 
        tier_formula = (
            f'=IF(J{r}>={tier1_threshold_cell},"Tier 1 (High)",'
            f'IF(J{r}>={tier2_threshold_cell},"Tier 2 (Medium)","Tier 3 (Low)"))'
        )
        tier_cell = ws.cell(row=r, column=11, value=tier_formula)
        tier_cell.font = FORMULA_FONT
        tier_cell.alignment = Alignment(horizontal="center", vertical="center")
 
        # Validation requirements via nested IF, mirroring tiering_engine.py exactly
        freq_formula = (
            f'=IF(K{r}="Tier 1 (High)","Annual",IF(K{r}="Tier 2 (Medium)","Every 18 months","Every 36 months"))'
        )
        mon_formula = (
            f'=IF(K{r}="Tier 1 (High)","Monthly",IF(K{r}="Tier 2 (Medium)","Quarterly","Semi-annual"))'
        )
        signoff_formula = (
            f'=IF(K{r}="Tier 1 (High)","Model Risk Committee",'
            f'IF(K{r}="Tier 2 (Medium)","Model Risk Management","Line of Business Risk Owner"))'
        )
        for col, formula in [(12, freq_formula), (13, mon_formula), (14, signoff_formula)]:
            cell = ws.cell(row=r, column=col, value=formula)
            cell.font = FORMULA_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
 
        rationale_cell = ws.cell(row=r, column=15, value=row_data["scoring_rationale"])
        rationale_cell.font = Font(name=ARIAL, size=9)
        rationale_cell.alignment = Alignment(wrap_text=True, vertical="top")
 
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 60
 
    last_row = header_row + len(df)
 
    # ---- Conditional formatting on Tier column ----
    tier_range = f"K{header_row+1}:K{last_row}"
    ws.conditional_formatting.add(tier_range, CellIsRule(operator="equal", formula=['"Tier 1 (High)"'], fill=TIER1_FILL))
    ws.conditional_formatting.add(tier_range, CellIsRule(operator="equal", formula=['"Tier 2 (Medium)"'], fill=TIER2_FILL))
    ws.conditional_formatting.add(tier_range, CellIsRule(operator="equal", formula=['"Tier 3 (Low)"'], fill=TIER3_FILL))
 
    # ---- Column widths ----
    widths = [10, 32, 28, 16, 18, 9, 9, 10, 10, 9, 14, 12, 11, 18, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions[get_column_letter(weight_start_col)].width = 22
    ws.column_dimensions[get_column_letter(weight_start_col+1)].width = 10
 
    ws.freeze_panes = "C5"
    ws.row_dimensions[header_row].height = 30
 
    # =====================================================================
    # SHEET 2: TIERING SUMMARY DASHBOARD
    # =====================================================================
    ws2 = wb.create_sheet("Tiering Summary")
    ws2["A1"] = "Model Inventory — Tiering Summary"
    ws2["A1"].font = TITLE_FONT
    ws2["A2"] = "Aggregated view for Model Risk Committee reporting"
    ws2["A2"].font = SUBTITLE_FONT
 
    ws2["A4"] = "Tier"
    ws2["B4"] = "Model Count"
    ws2["C4"] = "% of Inventory"
    ws2["D4"] = "Validation Frequency"
    ws2["E4"] = "Monitoring Frequency"
    style_header_row(ws2, 4, 5)
 
    tiers = ["Tier 1 (High)", "Tier 2 (Medium)", "Tier 3 (Low)"]
    fills = [TIER1_FILL, TIER2_FILL, TIER3_FILL]
    for i, tier in enumerate(tiers):
        r = 5 + i
        ws2.cell(row=r, column=1, value=tier).font = BODY_FONT
        count_formula = f"=COUNTIF('Model Inventory'!K{header_row+1}:K{last_row},\"{tier}\")"
        ws2.cell(row=r, column=2, value=count_formula).font = FORMULA_FONT
        pct_formula = f"=B{r}/SUM($B$5:$B$7)"
        pct_cell = ws2.cell(row=r, column=3, value=pct_formula)
        pct_cell.font = FORMULA_FONT
        pct_cell.number_format = "0%"
        req = VALIDATION_REQUIREMENTS[tier]
        ws2.cell(row=r, column=4, value=req["validation_frequency"]).font = BODY_FONT
        ws2.cell(row=r, column=5, value=req["monitoring_frequency"]).font = BODY_FONT
        for c in range(1, 6):
            ws2.cell(row=r, column=c).fill = fills[i]
            ws2.cell(row=r, column=c).border = BORDER
 
    ws2["A9"] = "Total models in inventory:"
    ws2["A9"].font = Font(name=ARIAL, bold=True, size=10)
    ws2["B9"] = "=SUM(B5:B7)"
    ws2["B9"].font = Font(name=ARIAL, bold=True, size=10)
 
    ws2["A11"] = "Validation Workload (next 12 months)"
    ws2["A11"].font = Font(name=ARIAL, bold=True, size=12, color="1F3864")
    ws2["A12"] = "Tier 1 models requiring validation THIS YEAR (annual cycle):"
    ws2["A12"].font = BODY_FONT
    ws2["B12"] = "=B5"
    ws2["B12"].font = FORMULA_FONT
    ws2["A13"] = "Tier 2 models requiring validation THIS YEAR (assume ~67% due, 18-month cycle):"
    ws2["A13"].font = BODY_FONT
    ws2["B13"] = "=ROUND(B6*0.67,0)"
    ws2["B13"].font = FORMULA_FONT
    ws2["A14"] = "Tier 3 models requiring validation THIS YEAR (assume ~33% due, 36-month cycle):"
    ws2["A14"].font = BODY_FONT
    ws2["B14"] = "=ROUND(B7*0.33,0)"
    ws2["B14"].font = FORMULA_FONT
    ws2["A15"] = "Estimated total validations due this year:"
    ws2["A15"].font = Font(name=ARIAL, bold=True, size=10)
    ws2["B15"] = "=B12+B13+B14"
    ws2["B15"].font = Font(name=ARIAL, bold=True, size=10)
 
    for col, w in [("A", 58), ("B", 14), ("C", 14), ("D", 20), ("E", 20)]:
        ws2.column_dimensions[col].width = w
 
    wb.save("inventory/Model_Inventory.xlsx")
    print(f"Saved inventory/Model_Inventory.xlsx with {len(df)} models.")
 
 
if __name__ == "__main__":
    build_workbook()